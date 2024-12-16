#Importing libraries
from tkinter import *
import tkinter as tk
from tkinter import ttk
from openpyxl import *
import numpy
import pygame
from time import strftime
from datetime import datetime
from time import sleep
from tkinter import Tk, Label, PhotoImage

#Initialization - (paths not updated as of July 30 2024) 
sound_path = "C:/Users/Micro/Desktop/MathLab GUI/Important/click.wav"
wallpaper_path = "C:/Users/Micro/Desktop/MathLab GUI/Important/main.png"
sheet_path = "C:/Users/Micro/Desktop/MathLab GUI/Important/sheet.xlsx"
ringtone_path = "C:/Users/Micro/Desktop/MathLab GUI/Important/ringtone.mp3"
pygame.mixer.init()
pathx, pathy, logox = 40, 30, 1160
for3x1, for3x2, for3x3, for3y =  90, 490, 890, 350
quadraticx, quadraticy = 240, 135
cubicx, cubicy = 310, 100
quarticx, quarticy = 310, 90
calculatorx, calculatory = 375, 30
converterx, convertery = 340, 145
milliseconds, running = 0, False
clockx, clocky = 270, 140
stopwatchx, stopwatchy = 500, 150
notepadx, notepady = 315, 150
formx = 400
running, times, minutes, hours, seconds, count = False, 0, 0, 0, 0, 1
i = "Number of "
 
#Main Functions
def design():
    global new_window
    new_window = tk.Toplevel(window)
    new_window.attributes('-fullscreen', True)
    new_window.configure(bg="#262626")

def logo():
    path = tk.Label(new_window, text = "MathLab GUI", bg="#7cd8e0")
    path.place(x = logox, y = pathy)

def return_button():
    #Buttons Color execution functions
    def back_enter(nul):
        back_button.config(bg="#ff8f03", width=13, height=1)

    def back_leave(nul):
        back_button.config(bg="SystemButtonFace", width=10, height=1)

    #Back button
    back_button = tk.Button(new_window, text = "Go Back ↩", command=new_window.destroy, activebackground="yellow", width=10, height=1, font = ("Consolas", 10))
    back_button.place(x=1165, y=660)

    #Buttons Color execution
    back_button.bind("<Enter>", back_enter)
    back_button.bind("<Leave>", back_leave)

def sound():
    click = pygame.mixer.Sound(sound_path)
    click.play()

#Polynomials' Applications
def quadratic():
    def quadratic_inner():
        design()
        logo()

        #Start
        path = tk.Label(new_window, text="MAIN → Applications → Math → Polynomials → Quadratic Equation", bg="#5be471")
        path.place(x = pathx, y = pathy)

        frame = tk.Frame(new_window, bg="#262626")
        frame.place(x = quadraticx, y = quadraticy)

        #Quadratic Equation code
        new_label = tk.Label(frame, text="Quadratic Equation Calculator", font= ("Consolas", 22), bg="#262626", fg="#ff99cc")
        new_label.grid(row=0, column=0, columnspan=4)

        equation_label = tk.Label(frame, text="ax² + bx + c = 0", font= ("Consolas", 22), bg="#262626", fg="#ff9999")
        equation_label.grid(row=1, column=0, columnspan=4)

        a_value = tk.Label(frame, text="Value of a: ", bg="#262626", fg="#ffff99", font=("Consolas", 15))
        b_value = tk.Label(frame, text="Value of b: ", bg="#262626", fg="#ffff99", font=("Consolas", 15))
        c_value = tk.Label(frame, text="Value of c: ", bg="#262626", fg="#ffff99", font=("Consolas", 15))

        a_input = tk.Entry(frame, width=40, borderwidth=2, fg="white", bg="#737373", font=("Consolas", 15))
        b_input = tk.Entry(frame, width=40, borderwidth=2, fg="white", bg="#737373", font=("Consolas", 15))
        c_input = tk.Entry(frame, width=40, borderwidth=2, fg="white", bg="#737373", font=("Consolas", 15))

        a_value.grid(row=2, column=0, padx=20, pady=20)
        b_value.grid(row=3, column=0, padx=20, pady=20)
        c_value.grid(row=4, column=0, padx=20, pady=20)

        a_input.grid(row=2, column=1, padx=100, pady=20)
        b_input.grid(row=3, column=1, padx=100, pady=20)
        c_input.grid(row=4, column=1, padx=100, pady=20)

        def quadratic_submit_button():
            global wrong_input
            global x_values
            global x
            global eq_label
            global equation
            a_num = a_input.get()
            b_num = b_input.get()
            c_num = c_input.get()
            try:
                wrong_input.destroy()
            except:
                pass
            try:
                x_values.destroy()
                x.destroy()
            except:
                pass
            try:
                eq_label.destroy()
                equation.destroy()
            except:
                pass
            try:
                a_float = float(a_num)
                b_float = float(b_num)
                c_float = float(c_num)
                ans = numpy.roots([a_float, b_float, c_float])
                ans_1 = numpy.round(ans[0], 3)
                ans_2 = numpy.round(ans[1], 3)
                x_values = tk.Label(frame, text="Roots of x: ", fg="#ff9999", bg="#262626", font=("Consolas", 15))
                x_values.grid(row=8, column=0, padx=20, pady=20)
                if ans_1 == ans_2:
                    x = tk.Label(frame, text=str(ans_1), fg="white", bg="#262626", font=("Consolas", 15))
                else:
                    x = tk.Label(frame, text=str(ans_1) + ", " + str(ans_2), fg="white", bg="#262626", font=("Consolas", 15))

                x.grid(row=8, column=1, padx=100, pady=20)
            except:
                wrong_input = tk.Label(frame, text="Error: Invalid Input", fg="#ff0000", bg="#262626", font=("Consolas", 15))
                wrong_input.grid(row=6, column=0, pady=20)

        def quadratic_clear_button():
            global wrong_input
            global x_values
            global x
            global eq_label
            global equation
            a_input.delete(0, "end")
            b_input.delete(0, "end")
            c_input.delete(0, "end")
            try:
                wrong_input.destroy()
            except:
                pass
            try:
                x_values.destroy()
                x.destroy()
            except:
                pass
            try:
                eq_label.destroy()
                equation.destroy()
            except:
                pass

        submit_button = tk.Button(frame, text="Submit", fg="#66FF99", bg="#737373", font=("Consolas", 12), command=quadratic_submit_button, width = 30)
        clear_button = tk.Button(frame, text="Clear", fg="#66FF99", bg="#737373", font=("Consolas", 12), command=quadratic_clear_button, width = 30)

        submit_button.grid(row=5, column=1, padx=100, pady=20)
        clear_button.grid(row=6, column=1, padx=100, pady=20)

        #End
        return_button()
    sound()
    quadratic_inner()

def cubic():
    def cubic_inner():
        design()
        logo()

        #Start
        path = tk.Label(new_window, text="MAIN → Applications → Math → Polynomials → Cubic Equation", bg="#5be471")
        path.place(x = pathx, y = pathy)
    

        #Cubic Equation code
        frame = tk.Frame(new_window, bg="#262626")
        frame.place(x = cubicx, y = cubicy)

        new_label = tk.Label(frame, text="Cubic Equation Calculator", font= ("Consolas", 22), bg="#262626", fg="#ff99cc")
        new_label.grid(row=0, column=0, columnspan=4)

        equation_label = tk.Label(frame, text="ax³ + bx² + cx + d = 0", font=("Consolas", 18), bg="#262626", fg="#ff9999")
        equation_label.grid(row=1, column=0, columnspan=4)

        a_value = tk.Label(frame, text="Value of a: ", bg="#262626", fg="#ffff99", font=("Consolas", 15))
        b_value = tk.Label(frame, text="Value of b: ", bg="#262626", fg="#ffff99", font=("Consolas", 15))
        c_value = tk.Label(frame, text="Value of c: ", bg="#262626", fg="#ffff99", font=("Consolas", 15))
        d_value = tk.Label(frame, text="Value of d: ", bg="#262626", fg="#ffff99", font=("Consolas", 15))

        a_input = tk.Entry(frame, width=40, borderwidth=2, fg="white", bg="#737373", font="Consolas")
        b_input = tk.Entry(frame, width=40, borderwidth=2, fg="white", bg="#737373", font="Consolas")
        c_input = tk.Entry(frame, width=40, borderwidth=2, fg="white", bg="#737373", font="Consolas")
        d_input = tk.Entry(frame, width=40, borderwidth=2, fg="white", bg="#737373", font="Consolas")

        a_value.grid(row=2, column=0, padx=20, pady=20)
        b_value.grid(row=3, column=0, padx=20, pady=20)
        c_value.grid(row=4, column=0, padx=20, pady=20)
        d_value.grid(row=5, column=0, padx=20, pady=20)

        a_input.grid(row=2, column=1, padx=100, pady=20)
        b_input.grid(row=3, column=1, padx=100, pady=20)
        c_input.grid(row=4, column=1, padx=100, pady=20)
        d_input.grid(row=5, column=1, padx=100, pady=20)

        def cubic_submit_button():
            global wrong_input
            global x_values
            global x
            global eq_label
            global equation
            a_num = a_input.get()
            b_num = b_input.get()
            c_num = c_input.get()
            d_num = d_input.get()
            try:
                wrong_input.destroy()
            except:
                pass
            try:
                x_values.destroy()
                x.destroy()
            except:
                pass
            try:
                eq_label.destroy()
                equation.destroy()
            except:
                pass
            try:
                a_float = float(a_num)
                b_float = float(b_num)
                c_float = float(c_num)
                d_float = float(d_num)
                ans = numpy.roots([a_float, b_float, c_float, d_float])
                ans_1 = numpy.round(ans[0], 3)
                ans_2 = numpy.round(ans[1], 3)
                ans_3 = numpy.round(ans[2], 3)
                x_values = tk.Label(frame, text="Roots of x: ", fg="#ff9999", bg="#262626",
                                    font=("Consolas", 15))
                x_values.grid(row=8, column=0, padx=20, pady=20)
                x = tk.Label(frame, text=str(ans_1) + ", " + str(ans_2) + ", " + str(ans_3), fg="white", bg="#262626",
                            font=("Consolas", 15))

                x.grid(row=8, column=1, padx=100, pady=20)
            except:
                wrong_input = tk.Label(frame, text="Error: Invalid Input", fg="#ff0000", bg="#262626", font=("Consolas", 15))
                wrong_input.grid(row=7, column=0, pady=20)

        def cubic_clear_button():
            global wrong_input
            global x_values
            global x
            global eq_label
            global equation
            a_input.delete(0, "end")
            b_input.delete(0, "end")
            c_input.delete(0, "end")
            d_input.delete(0, "end")
            try:
                wrong_input.destroy()
            except:
                pass
            try:
                x_values.destroy()
                x.destroy()
            except:
                pass
            try:
                eq_label.destroy()
                equation.destroy()
            except:
                pass

        submit_button = tk.Button(frame, text="Submit", fg="#66FF99", bg="#737373", font=("Consolas", 12), command=cubic_submit_button,
                                width=30)
        clear_button = tk.Button(frame, text="Clear", fg="#66FF99", bg="#737373", font=("Consolas", 12), command=cubic_clear_button,
                                width=30)

        submit_button.grid(row=6, column=1, padx=100, pady=20)
        clear_button.grid(row=7, column=1, padx=100, pady=20)

        #End
        return_button()
    sound()
    cubic_inner()

def quartic():
    def quartic_inner():
        design()
        logo()

        #Start
        path = tk.Label(new_window, text="MAIN → Applications → Math → Polynomials → Quartic Equation", bg="#5be471")
        path.place(x = pathx, y = pathy)
    

        #Quartic Equation code
        frame = tk.Frame(new_window, bg="#262626")
        frame.place(x = quarticx, y = quarticy)

        new_label = tk.Label(frame, text="Quartic Equation Calculator", font=("Consolas", 22), bg="#262626", fg="#ff99cc")
        new_label.grid(row=0, column=0, columnspan=4)

        equation_label = tk.Label(frame, text="ax⁴ + bx³ + cx² + dx + e = 0", font=("Consolas", 18), bg="#262626", fg="#ff9999")
        equation_label.grid(row=1, column=0, columnspan=4)

        a_value = tk.Label(frame, text="Value of a: ", bg="#262626", fg="#ffff99", font=("Consolas", 15))
        b_value = tk.Label(frame, text="Value of b: ", bg="#262626", fg="#ffff99", font=("Consolas", 15))
        c_value = tk.Label(frame, text="Value of c: ", bg="#262626", fg="#ffff99", font=("Consolas", 15))
        d_value = tk.Label(frame, text="Value of d: ", bg="#262626", fg="#ffff99", font=("Consolas", 15))
        e_value = tk.Label(frame, text="Value of e: ", bg="#262626", fg="#ffff99", font=("Consolas", 15))

        a_input = tk.Entry(frame, width=40, borderwidth=2, fg="white", bg="#737373", font="Consolas")
        b_input = tk.Entry(frame, width=40, borderwidth=2, fg="white", bg="#737373", font="Consolas")
        c_input = tk.Entry(frame, width=40, borderwidth=2, fg="white", bg="#737373", font="Consolas")
        d_input = tk.Entry(frame, width=40, borderwidth=2, fg="white", bg="#737373", font="Consolas")
        e_input = tk.Entry(frame, width=40, borderwidth=2, fg="white", bg="#737373", font="Consolas")

        a_value.grid(row=2, column=0, padx=20, pady=20)
        b_value.grid(row=3, column=0, padx=20, pady=20)
        c_value.grid(row=4, column=0, padx=20, pady=20)
        d_value.grid(row=5, column=0, padx=20, pady=20)
        e_value.grid(row=6, column=0, padx=20, pady=20)

        a_input.grid(row=2, column=1, padx=100, pady=20)
        b_input.grid(row=3, column=1, padx=100, pady=20)
        c_input.grid(row=4, column=1, padx=100, pady=20)
        d_input.grid(row=5, column=1, padx=100, pady=20)
        e_input.grid(row=6, column=1, padx=100, pady=20)

        def quartic_submit_button():
            global wrong_input
            global x_values
            global x
            global eq_label
            global equation
            a_num = a_input.get()
            b_num = b_input.get()
            c_num = c_input.get()
            d_num = d_input.get()
            e_num = e_input.get()
            try:
                wrong_input.destroy()
            except:
                pass
            try:
                x_values.destroy()
                x.destroy()
            except:
                pass
            try:
                eq_label.destroy()
                equation.destroy()
            except:
                pass
            try:
                a_float = float(a_num)
                b_float = float(b_num)
                c_float = float(c_num)
                d_float = float(d_num)
                e_float = float(e_num)
                ans = numpy.roots([a_float, b_float, c_float, d_float, e_float])
                ans_1 = numpy.round(ans[0], 3)
                ans_2 = numpy.round(ans[1], 3)
                ans_3 = numpy.round(ans[2], 3)
                ans_4 = numpy.round(ans[3], 3)
                x_values = tk.Label(frame, text="Roots of x: ", fg="#ff9999", bg="#262626", font=("Consolas", 15))
                x_values.grid(row=9, column=0, padx=20, pady=20)
                x = tk.Label(frame, text=str(ans_1) + ", " + str(ans_2) + ", " + str(ans_3) + ", " + str(ans_4), fg="white",
                            bg="#262626", font=("Consolas", 11))

                x.grid(row=9, column=1, padx=100, pady=20)
            except:
                wrong_input = tk.Label(frame, text="Error: Invalid Input", fg="#ff0000", bg="#262626", font=("Consolas", 15))
                wrong_input.grid(row=8, column=0, pady=20)

        def quartic_clear_button():
            global wrong_input
            global x_values
            global x
            global eq_label
            global equation
            a_input.delete(0, "end")
            b_input.delete(0, "end")
            c_input.delete(0, "end")
            d_input.delete(0, "end")
            e_input.delete(0, "end")
            try:
                wrong_input.destroy()
            except:
                pass
            try:
                x_values.destroy()
                x.destroy()
            except:
                pass
            try:
                eq_label.destroy()
                equation.destroy()
            except:
                pass

        submit_button = tk.Button(frame, text="Submit", fg="#66FF99", bg="#737373", font=("Consolas", 12), command=quartic_submit_button,
                                width=30)
        clear_button = tk.Button(frame, text="Clear", fg="#66FF99", bg="#737373", font=("Consolas", 12), command=quartic_clear_button,
                                width=30)

        submit_button.grid(row=7, column=1, padx=100, pady=20)
        clear_button.grid(row=8, column=1, padx=100, pady=20)

        #End
        return_button()
    sound()
    quartic_inner()

#Math Applications
def calculator():
    def calculator_inner():
        design()
        logo()

        #Start
        path = tk.Label(new_window, text="MAIN → Applications → Math → Calculator", bg="#5be471")
        path.place(x = pathx, y = pathy)

        frame = tk.Frame(new_window, pady=35, bg="#262626")
        frame.place(x = calculatorx, y = calculatory)


        new_label = tk.Label(frame, text="Calculator", font=("Consolas", 25), bg="#262626", fg="#ff99cc")
        new_label.grid(row=0, column=0, columnspan=4)

        #Calculator code
        num = tk.Entry(frame, width=30, borderwidth=5, font="Consolas", bg="#262626", fg="white")
        num.grid(row=1, column=0, columnspan=4, padx=20, pady=40)

        def button_click(number):
            current = num.get()
            num.delete(0, "end")
            num.insert(0, str(current) + str(number))

        def button_erase():
            length = num.get()
            num.delete(len(length) - 1)

        def button_clear():
            num.delete(0, "end")

        def button_add():
            global operation
            global f_num
            first_number = num.get()
            f_num = float(first_number)
            operation = "addition"
            num.delete(0, "end")

        def button_subtract():
            global operation
            global f_num
            first_number = num.get()
            f_num = float(first_number)
            operation = "subtraction"
            num.delete(0, "end")

        def button_multiply():
            global operation
            global f_num
            first_number = num.get()
            f_num = float(first_number)
            operation = "multiplication"
            num.delete(0, "end")

        def button_divide():
            global operation
            global f_num
            first_number = num.get()
            f_num = float(first_number)
            operation = "division"
            num.delete(0, "end")

        def button_equal():
            global operation
            global f_num
            second_number = num.get()
            num.delete(0, "end")
            if operation == "addition":
                num.insert(0, f_num + float(second_number))
            elif operation == "subtraction":
                num.insert(0, f_num - float(second_number))
            elif operation == "multiplication":
                num.insert(0, f_num * float(second_number))
            elif operation == "division":
                num.insert(0, f_num / float(second_number))

        button_point = tk.Button(frame, text=".", padx=53, pady=20, fg="white", bg="#4d4d4d", font="Consolas",
                                command=lambda: button_click("."))
        button_0 = tk.Button(frame, text="0", padx=61, pady=20, fg="white", bg="#4d4d4d", font="Consolas",
                            command=lambda: button_click(0))
        button_1 = tk.Button(frame, text="1", padx=61, pady=20, fg="white", bg="#4d4d4d", font="Consolas",
                            command=lambda: button_click(1))
        button_2 = tk.Button(frame, text="2", padx=53, pady=20, fg="white", bg="#4d4d4d", font="Consolas",
                            command=lambda: button_click(2))
        button_3 = tk.Button(frame, text="3", padx=51, pady=20, fg="white", bg="#4d4d4d", font="Consolas",
                            command=lambda: button_click(3))
        button_4 = tk.Button(frame, text="4", padx=61, pady=20, fg="white", bg="#4d4d4d", font="Consolas",
                            command=lambda: button_click(4))
        button_5 = tk.Button(frame, text="5", padx=53, pady=20, fg="white", bg="#4d4d4d", font="Consolas",
                            command=lambda: button_click(5))
        button_6 = tk.Button(frame, text="6", padx=51, pady=20, fg="white", bg="#4d4d4d", font="Consolas",
                            command=lambda: button_click(6))
        button_7 = tk.Button(frame, text="7", padx=61, pady=20, fg="white", bg="#4d4d4d", font="Consolas",
                            command=lambda: button_click(7))
        button_8 = tk.Button(frame, text="8", padx=53, pady=20, fg="white", bg="#4d4d4d", font="Consolas",
                            command=lambda: button_click(8))
        button_9 = tk.Button(frame, text="9", padx=51, pady=20, fg="white", bg="#4d4d4d", font="Consolas",
                            command=lambda: button_click(9))
        button_divide = tk.Button(frame, text="÷", padx=50, pady=20, fg="white", bg="#737373", font="Consolas",
                                command=button_divide)
        button_multiply = tk.Button(frame, text="×", padx=50, pady=20, fg="white", bg="#737373", font="Consolas",
                                    command=button_multiply)
        button_subtract = tk.Button(frame, text="-", padx=50, pady=20, fg="white", bg="#737373", font="Consolas",
                                    command=button_subtract)
        button_add = tk.Button(frame, text="+", padx=50, pady=20, fg="white", bg="#737373", font="Consolas",
                            command=button_add)
        button_equal = tk.Button(frame, text="=", padx=51, pady=20, fg="white", bg="#737373", font="Consolas",
                                command=button_equal)
        button_clear = tk.Button(frame, text="C", padx=125, pady=20, fg="white", bg="#737373", font="Consolas",
                                command=button_clear)
        button_erase = tk.Button(frame, text="⌫", padx=106, pady=20, fg="white", bg="#737373", font="Consolas",
                                command=button_erase)

        button_clear.grid(row=2, column=0, columnspan=2)
        button_erase.grid(row=2, column=2, columnspan=2)

        button_7.grid(row=3, column=0)
        button_8.grid(row=3, column=1)
        button_9.grid(row=3, column=2)
        button_divide.grid(row=3, column=3)

        button_4.grid(row=4, column=0)
        button_5.grid(row=4, column=1)
        button_6.grid(row=4, column=2)
        button_multiply.grid(row=4, column=3)

        button_1.grid(row=5, column=0)
        button_2.grid(row=5, column=1)
        button_3.grid(row=5, column=2)
        button_subtract.grid(row=5, column=3)

        button_0.grid(row=6, column=0)
        button_point.grid(row=6, column=1)
        button_equal.grid(row=6, column=2)
        button_add.grid(row=6, column=3)


        #End
        return_button()
    sound()
    calculator_inner()

def polynomials():
    def polynomials_inner():
        design()
        logo()

        #Start
        path = tk.Label(new_window, text="MAIN → Applications → Math → Polynomials", bg="#5be471")
        path.place(x = pathx, y = pathy)

        #Polynomials code
        heading = tk.Label(new_window, text="Polynomials", bg="#262626", fg="#99ccff", font=("Calibri", 25))
        heading.place(x = 553, y = pathy)
    
        #Buttons Color execution functions
        def quadratic_enter(nul):
            quadratic_button.config(bg="#46d149", width=32, height=3)

        def quadratic_leave(nul):
            quadratic_button.config(bg="SystemButtonFace", width=30, height=2)

        def cubic_enter(nul):
            cubic_button.config(bg="#46d149", width=32, height=3)

        def cubic_leave(nul):
            cubic_button.config(bg="SystemButtonFace", width=30, height=2)

        def quartic_enter(nul):
            quartic_button.config(bg="#46d149", width=32, height=3)

        def quartic_leave(nul):
            quartic_button.config(bg="SystemButtonFace", width=30, height=2)

        #Quadratic Equation button
        quadratic_button = tk.Button(new_window, text = "Quadratic Equation", command = quadratic, width = 30, height = 2, font = "Consolas", activebackground="pink")
        quadratic_button.place(x = for3x1, y = for3y)

        #Cubic Equation button
        cubic_button = tk.Button(new_window, text = "Cubic Equation", command = cubic, width = 30, height = 2, font = "Consolas", activebackground="pink")
        cubic_button.place(x = for3x2, y = for3y)

        #Quartic Equation button
        quartic_button = tk.Button(new_window, text = "Quartic Equation", command = quartic, width = 30, height = 2, font = "Consolas", activebackground="pink")
        quartic_button.place(x = for3x3, y = for3y)

        #Buttons Color execution
        quadratic_button.bind("<Enter>", quadratic_enter)
        quadratic_button.bind("<Leave>", quadratic_leave)

        cubic_button.bind("<Enter>", cubic_enter)
        cubic_button.bind("<Leave>", cubic_leave)

        quartic_button.bind("<Enter>", quartic_enter)
        quartic_button.bind("<Leave>", quartic_leave)

        #End
        return_button()
    sound()
    polynomials_inner()

def conversions():
    def conversions_inner():
        design()
        logo()
        #Start
        path = tk.Label(new_window, text="MAIN → Applications → Math → Digital Storage Converter", bg="#5be471")
        path.place(x = pathx, y = pathy)

        #Digial storage conversions code

        frame = tk.Frame(new_window, bg="#262626")
        frame.place(x = converterx, y = convertery)

        def selectedOption1(event):
            selected1 = drop1.get()
            selected2 = drop2.get()
            if selected1 == selected2:
                drop2.delete(0, "end")

        def selectedOption2(event):
            selected1 = drop1.get()
            selected2 = drop2.get()
            if selected1 == selected2:
                drop1.delete(0, "end")

        def conversion():
            selected1 = drop1.get()
            selected2 = drop2.get()
            if selected1 == "Bit" or selected2 == "Bit":
                bit()
            if selected1 == "Byte" or selected2 == "Byte":
                byte()
            if selected1 == "Kilobyte" or selected2 == "Kilobyte":
                kilobyte()
            if selected1 == "Megabyte" or selected2 == "Megabyte":
                megabyte()
            if selected1 == "Gigabyte" or selected2 == "Gigabyte":
                gigabyte()
            if selected1 == "Terabyte" or selected2 == "Terabyte":
                terabyte()

        def bit():
            selected1 = drop1.get()
            selected2 = drop2.get()
            num1 = unit_1.get()
            num2 = unit_2.get()

            if selected1 == "Byte":
                try:
                    num2 = float(num2)
                except:
                    return
                try:
                    unit_1.delete(0, "end")
                except:
                    pass
                ans = num2 / 8
                unit_1.insert(0, str(ans))
            if selected2 == "Byte":
                try:
                    num1 = float(num1)
                except:
                    return
                try:
                    unit_2.delete(0, "end")
                except:
                    pass
                ans = num1 / 8
                unit_2.insert(0, str(ans))

            if selected1 == "Kilobyte":
                try:
                    num2 = float(num2)
                except:
                    return
                try:
                    unit_1.delete(0, "end")
                except:
                    pass
                ans = num2 / (8 * 1000)
                unit_1.insert(0, str(ans))
            if selected2 == "Kilobyte":
                try:
                    num1 = float(num1)
                except:
                    return
                try:
                    unit_2.delete(0, "end")
                except:
                    pass
                ans = num1 / (8 * 1000)
                unit_2.insert(0, str(ans))

            if selected1 == "Megabyte":
                try:
                    num2 = float(num2)
                except:
                    return
                try:
                    unit_1.delete(0, "end")
                except:
                    pass
                ans = num2 / (8 * 1000 * 1000)
                unit_1.insert(0, str(ans))
            if selected2 == "Megabyte":
                try:
                    num1 = float(num1)
                except:
                    return
                try:
                    unit_2.delete(0, "end")
                except:
                    pass
                ans = num1 / (8 * 1000 * 1000)
                unit_2.insert(0, str(ans))

            if selected1 == "Gigabyte":
                try:
                    num2 = float(num2)
                except:
                    return
                try:
                    unit_1.delete(0, "end")
                except:
                    pass
                ans = num2 / (8 * 1000 * 1000 * 1000)
                unit_1.insert(0, str(ans))
            if selected2 == "Gigabyte":
                try:
                    num1 = float(num1)
                except:
                    return
                try:
                    unit_2.delete(0, "end")
                except:
                    pass
                ans = num1 / (8 * 1000 * 1000 * 1000)
                unit_2.insert(0, str(ans))

            if selected1 == "Terabyte":
                try:
                    num2 = float(num2)
                except:
                    return
                try:
                    unit_1.delete(0, "end")
                except:
                    pass
                ans = num2 / (8 * 1000 * 1000 * 1000 * 1000)
                unit_1.insert(0, str(ans))
            if selected2 == "Terabyte":
                try:
                    num1 = float(num1)
                except:
                    return
                try:
                    unit_2.delete(0, "end")
                except:
                    pass
                ans = num1 / (8 * 1000 * 1000 * 1000 * 1000)
                unit_2.insert(0, str(ans))

        def byte():
            selected1 = drop1.get()
            selected2 = drop2.get()
            num1 = unit_1.get()
            num2 = unit_2.get()

            if selected1 == "Bit":
                try:
                    num2 = float(num2)
                except:
                    return
                try:
                    unit_1.delete(0, "end")
                except:
                    pass
                ans = num2 * 8
                unit_1.insert(0, str(ans))
            if selected2 == "Bit":
                try:
                    num1 = float(num1)
                except:
                    return
                try:
                    unit_2.delete(0, "end")
                except:
                    pass
                ans = num1 * 8
                unit_2.insert(0, str(ans))

            if selected1 == "Kilobyte":
                try:
                    num2 = float(num2)
                except:
                    return
                try:
                    unit_1.delete(0, "end")
                except:
                    pass
                ans = num2 / 1000
                unit_1.insert(0, str(ans))
            if selected2 == "Kilobyte":
                try:
                    num1 = float(num1)
                except:
                    return
                try:
                    unit_2.delete(0, "end")
                except:
                    pass
                ans = num1 / 1000
                unit_2.insert(0, str(ans))

            if selected1 == "Megabyte":
                try:
                    num2 = float(num2)
                except:
                    return
                try:
                    unit_1.delete(0, "end")
                except:
                    pass
                ans = num2 / (1000 * 1000)
                unit_1.insert(0, str(ans))
            if selected2 == "Megabyte":
                try:
                    num1 = float(num1)
                except:
                    return
                try:
                    unit_2.delete(0, "end")
                except:
                    pass
                ans = num1 / (1000 * 1000)
                unit_2.insert(0, str(ans))

            if selected1 == "Gigabyte":
                try:
                    num2 = float(num2)
                except:
                    return
                try:
                    unit_1.delete(0, "end")
                except:
                    pass
                ans = num2 / (1000 * 1000 * 1000)
                unit_1.insert(0, str(ans))
            if selected2 == "Gigabyte":
                try:
                    num1 = float(num1)
                except:
                    return
                try:
                    unit_2.delete(0, "end")
                except:
                    pass
                ans = num1 / (1000 * 1000 * 1000)
                unit_2.insert(0, str(ans))

            if selected1 == "Terabyte":
                try:
                    num2 = float(num2)
                except:
                    return
                try:
                    unit_1.delete(0, "end")
                except:
                    pass
                ans = num2 / (1000 * 1000 * 1000 * 1000)
                unit_1.insert(0, str(ans))
            if selected2 == "Terabyte":
                try:
                    num1 = float(num1)
                except:
                    return
                try:
                    unit_2.delete(0, "end")
                except:
                    pass
                ans = num1 / (1000 * 1000 * 1000 * 1000)
                unit_2.insert(0, str(ans))

        def kilobyte():
            selected1 = drop1.get()
            selected2 = drop2.get()
            num1 = unit_1.get()
            num2 = unit_2.get()

            if selected1 == "Bit":
                try:
                    num2 = float(num2)
                except:
                    return
                try:
                    unit_1.delete(0, "end")
                except:
                    pass
                ans = num2 * (8 * 1000)
                unit_1.insert(0, str(ans))
            if selected2 == "Bit":
                try:
                    num1 = float(num1)
                except:
                    return
                try:
                    unit_2.delete(0, "end")
                except:
                    pass
                ans = num1 * (8 * 1000)
                unit_2.insert(0, str(ans))

            if selected1 == "Byte":
                try:
                    num2 = float(num2)
                except:
                    return
                try:
                    unit_1.delete(0, "end")
                except:
                    pass
                ans = num2 * 1000
                unit_1.insert(0, str(ans))
            if selected2 == "Byte":
                try:
                    num1 = float(num1)
                except:
                    return
                try:
                    unit_2.delete(0, "end")
                except:
                    pass
                ans = num1 * 1000
                unit_2.insert(0, str(ans))

            if selected1 == "Megabyte":
                try:
                    num2 = float(num2)
                except:
                    return
                try:
                    unit_1.delete(0, "end")
                except:
                    pass
                ans = num2 / 1000
                unit_1.insert(0, str(ans))
            if selected2 == "Megabyte":
                try:
                    num1 = float(num1)
                except:
                    return
                try:
                    unit_2.delete(0, "end")
                except:
                    pass
                ans = num1 / 1000
                unit_2.insert(0, str(ans))

            if selected1 == "Gigabyte":
                try:
                    num2 = float(num2)
                except:
                    return
                try:
                    unit_1.delete(0, "end")
                except:
                    pass
                ans = num2 / (1000 * 1000)
                unit_1.insert(0, str(ans))
            if selected2 == "Gigabyte":
                try:
                    num1 = float(num1)
                except:
                    return
                try:
                    unit_2.delete(0, "end")
                except:
                    pass
                ans = num1 / (1000 * 1000)
                unit_2.insert(0, str(ans))

            if selected1 == "Terabyte":
                try:
                    num2 = float(num2)
                except:
                    return
                try:
                    unit_1.delete(0, "end")
                except:
                    pass
                ans = num2 / (1000 * 1000 * 1000)
                unit_1.insert(0, str(ans))
            if selected2 == "Terabyte":
                try:
                    num1 = float(num1)
                except:
                    return
                try:
                    unit_2.delete(0, "end")
                except:
                    pass
                ans = num1 / (1000 * 1000 * 1000)
                unit_2.insert(0, str(ans))

        def megabyte():
            selected1 = drop1.get()
            selected2 = drop2.get()
            num1 = unit_1.get()
            num2 = unit_2.get()

            if selected1 == "Bit":
                try:
                    num2 = float(num2)
                except:
                    return
                try:
                    unit_1.delete(0, "end")
                except:
                    pass
                ans = num2 * (8 * 1000 * 1000)
                unit_1.insert(0, str(ans))
            if selected2 == "Bit":
                try:
                    num1 = float(num1)
                except:
                    return
                try:
                    unit_2.delete(0, "end")
                except:
                    pass
                ans = num1 * (8 * 1000 * 1000)
                unit_2.insert(0, str(ans))

            if selected1 == "Byte":
                try:
                    num2 = float(num2)
                except:
                    return
                try:
                    unit_1.delete(0, "end")
                except:
                    pass
                ans = num2 * (1000 * 1000)
                unit_1.insert(0, str(ans))
            if selected2 == "Byte":
                try:
                    num1 = float(num1)
                except:
                    return
                try:
                    unit_2.delete(0, "end")
                except:
                    pass
                ans = num1 * (1000 * 1000)
                unit_2.insert(0, str(ans))

            if selected1 == "Kilobyte":
                try:
                    num2 = float(num2)
                except:
                    return
                try:
                    unit_1.delete(0, "end")
                except:
                    pass
                ans = num2 * 1000
                unit_1.insert(0, str(ans))
            if selected2 == "Kilobyte":
                try:
                    num1 = float(num1)
                except:
                    return
                try:
                    unit_2.delete(0, "end")
                except:
                    pass
                ans = num1 * 1000
                unit_2.insert(0, str(ans))

            if selected1 == "Gigabyte":
                try:
                    num2 = float(num2)
                except:
                    return
                try:
                    unit_1.delete(0, "end")
                except:
                    pass
                ans = num2 / 1000
                unit_1.insert(0, str(ans))
            if selected2 == "Gigabyte":
                try:
                    num1 = float(num1)
                except:
                    return
                try:
                    unit_2.delete(0, "end")
                except:
                    pass
                ans = num1 / 1000
                unit_2.insert(0, str(ans))

            if selected1 == "Terabyte":
                try:
                    num2 = float(num2)
                except:
                    return
                try:
                    unit_1.delete(0, "end")
                except:
                    pass
                ans = num2 / (1000 * 1000)
                unit_1.insert(0, str(ans))
            if selected2 == "Terabyte":
                try:
                    num1 = float(num1)
                except:
                    return
                try:
                    unit_2.delete(0, "end")
                except:
                    pass
                ans = num1 / (1000 * 1000)
                unit_2.insert(0, str(ans))

        def gigabyte():
            selected1 = drop1.get()
            selected2 = drop2.get()
            num1 = unit_1.get()
            num2 = unit_2.get()

            if selected1 == "Bit":
                try:
                    num2 = float(num2)
                except:
                    return
                try:
                    unit_1.delete(0, "end")
                except:
                    pass
                ans = num2 * (8 * 1000 * 1000 * 1000)
                unit_1.insert(0, str(ans))
            if selected2 == "Bit":
                try:
                    num1 = float(num1)
                except:
                    return
                try:
                    unit_2.delete(0, "end")
                except:
                    pass
                ans = num1 * (8 * 1000 * 1000 * 1000)
                unit_2.insert(0, str(ans))

            if selected1 == "Byte":
                try:
                    num2 = float(num2)
                except:
                    return
                try:
                    unit_1.delete(0, "end")
                except:
                    pass
                ans = num2 * (1000 * 1000 * 1000)
                unit_1.insert(0, str(ans))
            if selected2 == "Byte":
                try:
                    num1 = float(num1)
                except:
                    return
                try:
                    unit_2.delete(0, "end")
                except:
                    pass
                ans = num1 * (1000 * 1000 * 1000)
                unit_2.insert(0, str(ans))

            if selected1 == "Kilobyte":
                try:
                    num2 = float(num2)
                except:
                    return
                try:
                    unit_1.delete(0, "end")
                except:
                    pass
                ans = num2 * (1000 * 1000)
                unit_1.insert(0, str(ans))
            if selected2 == "Kilobyte":
                try:
                    num1 = float(num1)
                except:
                    return
                try:
                    unit_2.delete(0, "end")
                except:
                    pass
                ans = num1 * (1000 * 1000)
                unit_2.insert(0, str(ans))

            if selected1 == "Megabyte":
                try:
                    num2 = float(num2)
                except:
                    return
                try:
                    unit_1.delete(0, "end")
                except:
                    pass
                ans = num2 * 1000
                unit_1.insert(0, str(ans))
            if selected2 == "Megabyte":
                try:
                    num1 = float(num1)
                except:
                    return
                try:
                    unit_2.delete(0, "end")
                except:
                    pass
                ans = num1 * 1000
                unit_2.insert(0, str(ans))

            if selected1 == "Terabyte":
                try:
                    num2 = float(num2)
                except:
                    return
                try:
                    unit_1.delete(0, "end")
                except:
                    pass
                ans = num2 / 1000
                unit_1.insert(0, str(ans))
            if selected2 == "Terabyte":
                try:
                    num1 = float(num1)
                except:
                    return
                try:
                    unit_2.delete(0, "end")
                except:
                    pass
                ans = num1 / 1000
                unit_2.insert(0, str(ans))

        def terabyte():
            selected1 = drop1.get()
            selected2 = drop2.get()
            num1 = unit_1.get()
            num2 = unit_2.get()

            if selected1 == "Bit":
                try:
                    num2 = float(num2)
                except:
                    return
                try:
                    unit_1.delete(0, "end")
                except:
                    pass
                ans = num2 * (8 * 1000 * 1000 * 1000 * 1000)
                unit_1.insert(0, str(ans))
            if selected2 == "Bit":
                try:
                    num1 = float(num1)
                except:
                    return
                try:
                    unit_2.delete(0, "end")
                except:
                    pass
                ans = num1 * (8 * 1000 * 1000 * 1000 * 1000)
                unit_2.insert(0, str(ans))

            if selected1 == "Byte":
                try:
                    num2 = float(num2)
                except:
                    return
                try:
                    unit_1.delete(0, "end")
                except:
                    pass
                ans = num2 * (1000 * 1000 * 1000 * 1000)
                unit_1.insert(0, str(ans))
            if selected2 == "Byte":
                try:
                    num1 = float(num1)
                except:
                    return
                try:
                    unit_2.delete(0, "end")
                except:
                    pass
                ans = num1 * (1000 * 1000 * 1000 * 1000)
                unit_2.insert(0, str(ans))

            if selected1 == "Kilobyte":
                try:
                    num2 = float(num2)
                except:
                    return
                try:
                    unit_1.delete(0, "end")
                except:
                    pass
                ans = num2 * (1000 * 1000 * 1000)
                unit_1.insert(0, str(ans))
            if selected2 == "Kilobyte":
                try:
                    num1 = float(num1)
                except:
                    return
                try:
                    unit_2.delete(0, "end")
                except:
                    pass
                ans = num1 * (1000 * 1000 * 1000)
                unit_2.insert(0, str(ans))

            if selected1 == "Megabyte":
                try:
                    num2 = float(num2)
                except:
                    return
                try:
                    unit_1.delete(0, "end")
                except:
                    pass
                ans = num2 * (1000 * 1000)
                unit_1.insert(0, str(ans))
            if selected2 == "Megabyte":
                try:
                    num1 = float(num1)
                except:
                    return
                try:
                    unit_2.delete(0, "end")
                except:
                    pass
                ans = num1 * (1000 * 1000)
                unit_2.insert(0, str(ans))

            if selected1 == "Gigabyte":
                try:
                    num2 = float(num2)
                except:
                    return
                try:
                    unit_1.delete(0, "end")
                except:
                    pass
                ans = num2 * 1000
                unit_1.insert(0, str(ans))
            if selected2 == "Gigabyte":
                try:
                    num1 = float(num1)
                except:
                    return
                try:
                    unit_2.delete(0, "end")
                except:
                    pass
                ans = num1 * 1000
                unit_2.insert(0, str(ans))

        def clear_converter():
            unit_1.delete(0, "end")
            unit_2.delete(0, "end")

        new_label = tk.Label(frame, text="Digital Storage Converter", font=("Consolas", 27), bg="#262626", fg="#ff99cc")
        unit_1 = tk.Entry(frame, width=20, borderwidth=2, fg="white", bg="#737373", font=("Consolas", 15))
        equal_label = tk.Label(frame, text="=", fg="white", bg="#262626", font=("Consolas", 30))
        unit_2 = tk.Entry(frame, width=20, borderwidth=2, fg="white", bg="#737373", font=("Consolas", 15))

        new_label.grid(row=0, column=0, columnspan=4)
        unit_1.grid(row=1, column=0, pady=20, padx=30)
        equal_label.grid(row=1, column=1, pady=20)
        unit_2.grid(row=1, column=2, pady=20, padx=30)

        options_1 = [
            "Bit",
            "Byte",
            "Kilobyte",
            "Megabyte",
            "Gigabyte",
            "Terabyte"
        ]

        options_2 = [
            "Bit",
            "Byte",
            "Kilobyte",
            "Megabyte",
            "Gigabyte",
            "Terabyte"
        ]

        drop1 = ttk.Combobox(frame, values=options_1, width=20, font=("Consolas", 14))
        drop2 = ttk.Combobox(frame, values=options_2, width=20, font=("Consolas", 14))

        drop1.set("Select")
        drop1.bind("<<ComboboxSelected>>", selectedOption1)

        drop2.set("Select")
        drop2.bind("<<ComboboxSelected>>", selectedOption2)

        convert_button = tk.Button(frame, text="Convert", fg="white", bg="#737373", font="Consolas", command=conversion)
        clear_button = tk.Button(frame, text="Clear", fg="white", bg="#737373", font="Consolas", command=clear_converter)

        drop1.grid(row=2, column=0)
        drop2.grid(row=2, column=2)
        convert_button.grid(row=3, column=0, columnspan=4, pady=50)
        clear_button.grid(row=4, column=0, columnspan=4)

        #End
        return_button()
    sound()
    conversions_inner()

#Time Applications
def clock():
    def clock_inner():
        design()
        logo()

        #Start
        path = tk.Label(new_window, text="MAIN → Applications → Time → Clock", bg="#5be471")
        path.place(x = pathx, y = pathy)

        #Clock code
        frame = tk.Frame(new_window, bg="#262626")
        frame.place(x = clockx, y = clocky)

        def clock_time():
            string = strftime('%H:%M:%S %p')
            lbl.config(text=string)
            lbl.after(1000, clock_time)

        new_label = tk.Label(frame, text="Digital Clock", font=("Consolas", 35), bg="#262626", fg="#ff99cc")
        lbl = tk.Label(frame, font=('calibri', 100, 'bold'),
                       foreground='white', bg="#262626")

        new_label.grid(row=0, column=0, columnspan=4, pady=20)
        lbl.grid(row=1, column=0, pady=20, padx=30)
        clock_time()

        #End
        return_button()
    sound()
    clock_inner()

def timer():
    def timer_inner():
        design()
        logo()

        #Start
        path = tk.Label(new_window, text="MAIN → Applications → Time → Timer", bg="#5be471")
        path.place(x = pathx, y = pathy)

        #Timer code
        new_window.option_add("*insertBackground", "#262626")

        frame = tk.Frame(new_window, bg="#262626")
        frame.place(x=stopwatchx, y=stopwatchy)

        def start():
            global times
            global seconds
            global hours
            global minutes
            global count
            global running
            running = True
            times = int(hrs.get()) * 3600 + int(mins.get()) * 60 + int(sec.get())
            while times > -1:
                if running is False:
                    break
                else:
                    minutes, seconds = (times // 60, times % 60)

                    if minutes >= 60:
                        hours, minutes = (minutes // 60, minutes % 60)
                        if len(str(hours)) > 1:
                            hrs.set(str(hours))
                        else:
                            hrs.set("0" + str(hours))

                    if minutes > 0:
                        if len(str(minutes)) > 1:
                            mins.set(str(minutes))
                        else:
                            mins.set("0" + str(minutes))
                    else:
                        mins.set("00")

                    if seconds > 0:
                        if len(str(seconds)) > 1:
                            sec.set(str(seconds))
                        else:
                            sec.set("0" + str(seconds))
                    else:
                        sec.set("00")

                    new_window.update()
                    sleep(1)

                    if times == 0:
                        ringtone = pygame.mixer.Sound(ringtone_path)
                        ringtone.play()

                    count += 1
                    times -= 1

        def stop():
            global times
            global seconds
            global hours
            global minutes
            global count
            global running
            running = False
            count -= 1

        def reset():
            global times
            global count
            global running
            if running is False:
                running = True
                times = int(hrs.get()) * 3600 + int(mins.get()) * 60 + int(sec.get()) + count - 1
                count = 1
                while times > -1:
                    if running is False:
                        break
                    else:
                        minutes, seconds = (times // 60, times % 60)

                        if minutes >= 60:
                            hours, minutes = (minutes // 60, minutes % 60)
                            if len(str(hours)) > 1:
                                hrs.set(str(hours))
                            else:
                                hrs.set("0" + str(hours))

                        if minutes > 0:
                            if len(str(minutes)) > 1:
                                mins.set(str(minutes))
                            else:
                                mins.set("0" + str(minutes))
                        else:
                            mins.set("00")

                        if seconds > 0:
                            if len(str(seconds)) > 1:
                                sec.set(str(seconds))
                            else:
                                sec.set("0" + str(seconds))
                        else:
                            sec.set("00")

                        new_window.update()
                        sleep(1)

                        if times == 0:
                            ringtone = pygame.mixer.Sound(ringtone_path)
                            ringtone.play()

                        count += 1
                        times -= 1
            else:
                times = int(hrs.get()) * 3600 + int(mins.get()) * 60 + int(sec.get()) + count
                count = 0

        new_label = tk.Label(frame, text="Timer", font=("Consolas", 22), bg="#262626", fg="#ff99cc")

        colon_label1 = tk.Label(frame, text=":", font=("Consolas", 22), bg="#262626", fg="white")
        colon_label2 = tk.Label(frame, text=":", font=("Consolas", 22), bg="#262626", fg="white")

        hrs = tk.StringVar()
        hrs_input = tk.Entry(frame, textvariable=hrs, bg="#262626", fg="white", font=("Consolas", 22), width=2, bd=0)
        hrs.set("00")

        mins = tk.StringVar()
        min_input = tk.Entry(frame, textvariable=mins, bg="#262626", fg="white", font=("Consolas", 22), width=2, bd=0)
        mins.set("00")

        sec = tk.StringVar()
        sec_input = tk.Entry(frame, textvariable=sec, bg="#262626", fg="white", font=("Consolas", 22), width=2, bd=0)
        sec.set("00")

        start_button = tk.Button(frame, text="▶", fg="white", bg="#737373", font=("Consolas", 22), command=start)
        stop_button = tk.Button(frame, text="⏸", fg="white", bg="#737373", font=("Consolas", 22), command=stop)
        reset_button = tk.Button(frame, text="↻", fg="white", bg="#737373", font=("Consolas", 22), command=reset)

        new_label.grid(pady=50, columnspan=5)
        hrs_input.grid(row=1, column=0, padx=10)
        colon_label1.grid(row=1, column=1)
        min_input.grid(row=1, column=2, padx=10)
        colon_label2.grid(row=1, column=3)
        sec_input.grid(row=1, column=4, padx=10)
        start_button.grid(row=2, column=0, pady=40, ipadx=20)
        stop_button.grid(row=2, column=2, ipadx=15)
        reset_button.grid(row=2, column=4, ipadx=18)

        #End
        return_button()
    sound()
    timer_inner()

def stopwatch():
    def stopwatch_inner():
        design()
        logo()

        #Start
        path = tk.Label(new_window, text="MAIN → Applications → Time → Stopwatch", bg="#5be471")
        path.place(x = pathx, y = pathy)

        #Stopwatch code
        frame = tk.Frame(new_window, bg="#262626")
        frame.place(x = stopwatchx, y = stopwatchy)


        def milliseconds_label(lab):
            def count():
                if running:
                    global milliseconds
                    tt = datetime.fromtimestamp(milliseconds)
                    if milliseconds > 600:
                        date_object = tt.strftime('%H:%M:%S')
                    elif milliseconds > 60:
                        date_object = tt.strftime('%M:%S:%f')[:-4]
                    else:
                        date_object = tt.strftime('%M:%S:%f')[:-4]
                    display = date_object

                    lab['text'] = display

                    lab.after(10, count)
                    milliseconds += 0.01

            count()

        def Start(label):
            global running
            running = True
            outside_timer = False
            milliseconds_label(label)

            start['state'] = 'disabled'
            start['highlightbackground'] = 'black'

            stop['state'] = 'normal'
            stop['highlightbackground'] = 'red'

            reset['state'] = 'normal'
            reset['highlightbackground'] = 'dark grey'

        def Stop():
            global running
            start['state'] = 'normal'
            start['highlightbackground'] = 'green'

            stop['state'] = 'disabled'
            stop['highlightbackground'] = 'black'

            reset['state'] = 'normal'
            reset['highlightbackground'] = 'dark grey'

            running = False

        def Reset(label):
            global milliseconds
            milliseconds = 00000
            if running == False:
                reset['state'] = 'disabled'
                reset['highlightbackground'] = 'black'

            label['text'] = "00:00:00"

        new_label = tk.Label(frame, text="Stopwatch", font=("Consolas", 30), bg="#262626", fg="#ff99cc")
        new_label.pack(pady=25)
        label = Label(frame, text="00:00:00", fg="white", font="Geneva 45 bold", bg="#262626")
        label.pack()

        start = Button(frame, text='Start', width=7, height=3, command=lambda: Start(label), font="Geneva 15 bold", fg="white", bg="#737373")
        reset = Button(frame, text='Reset', width=7, height=3, state='disabled', command=lambda: Reset(label), font="Geneva 15 bold", borderwidth=0, fg="white", bg="#737373")
        stop = Button(frame, text='Stop', width=7, height=3, state='disabled', command=Stop, font="Geneva 15 bold", fg="white", bg="#737373")


        start.pack(side="left", pady=40)
        reset.pack(side="left", pady=40)
        stop.pack(side="left", pady=40)


        #End
        return_button()
    sound()
    stopwatch_inner()

#Other Applications
def notepad():
    def notepad_inner():
        logo()
        new_notepad_window = tk.Toplevel(window)
        new_notepad_window.attributes('-fullscreen', True)
        new_notepad_window.configure(bg='#262626')

        #Start
        path = tk.Label(new_notepad_window, text="MAIN → Applications → Other → Notepad", bg="#5be471")
        path.place(x = pathx, y = pathy)

        notepad_logo = tk.Label(new_notepad_window, text="MathLab GUI", bg="#7cd8e0")
        notepad_logo.place(x = logox, y = pathy)

        #Notepad code
        heading = tk.Label(new_notepad_window, text="Notepad", bg="#262626", fg="#ff99cc", font=("Calibri", 25))
        heading.place(x = 577, y = pathy)

        notepad = tk.Text(new_notepad_window)
        notepad.place(x = notepadx, y = notepady)

        #Notepad Exit Button Color execution functions
        def notepad_exit_enter(nul):
            notepad_exit_button.config(bg="#e05122", width=17, height=1)

        def notepad_exit_leave(nul):
            notepad_exit_button.config(bg="SystemButtonFace", width=15, height=1)

        #Notepad Exit button
        notepad_exit_button = tk.Button(new_notepad_window, text = "Close Notepad", command = new_notepad_window.destroy, activebackground="red", width=15, height=1, font = ("Consolas", 10))
        notepad_exit_button.place(x = 1130, y = 660)

        #Notepad Button Color execution
        notepad_exit_button.bind("<Enter>", notepad_exit_enter)
        notepad_exit_button.bind("<Leave>", notepad_exit_leave)

        #End
        return_button()
    sound()
    notepad_inner()
    
def form():
    def form_inner():
        design()
        logo()

        #Start
        path = tk.Label(new_window, text="MAIN → Applications → Other → Student Form", bg="#5be471")
        path.place(x = pathx, y = pathy)

        #Form code
        heading = tk.Label(new_window, text="Student Form", bg="#262626", fg="#ff99cc", font=("Calibri", 25))
        heading.place(x = 540, y = pathy)

        wb = load_workbook(sheet_path)
        sheet = wb.active

        #Initialize excel sheet
        def excel():
            sheet.column_dimensions['A'].width = 15
            sheet.column_dimensions['B'].width = 13
            sheet.column_dimensions['C'].width = 12
            sheet.column_dimensions['D'].width = 17
            sheet.column_dimensions['E'].width = 17
            sheet.column_dimensions['F'].width = 20
            sheet.column_dimensions['G'].width = 18

            sheet.cell(row=1, column=1).value = "Name"
            sheet.cell(row=1, column=2).value = "Student ID"
            sheet.cell(row=1, column=3).value = "Grade"
            sheet.cell(row=1, column=4).value = "Section"
            sheet.cell(row=1, column=5).value = "Phone"
            sheet.cell(row=1, column=6).value = "Email"
            sheet.cell(row=1, column=7).value = "Address"

        #Refreshing
        def clear_form():
            name_field.delete(0, END)
            student_id_field.delete(0, END)
            grade_field.delete(0, END)
            section_field.delete(0, END)
            phone_field.delete(0, END)
            email_field.delete(0, END)
            address_field.delete(0, END)

        #Data In
        def insert():
            if (name_field.get() == "" and student_id_field.get() == "" and grade_field.get() == "" and section_field.get() == "" and
                phone_field.get() == "" and email_field.get() == "" and address_field.get() == ""):

                def message():
                    empty_promt['text'] = ("")
                    empty_promt['background'] = "#262626"


                def delay():
                    empty_promt['text'] = "No data entered, fill in the information"
                    empty_promt['font'] = "Calibri"
                    empty_promt['background'] = "#262626"
                    empty_promt['foreground'] = "light green"
                    new_window.after(1000, message)

                empty_promt = Label(new_window)
                empty_promt.place(x = formx + 100, y = 500)

            else:
                current_row = sheet.max_row

                sheet.cell(row=current_row + 1, column=1).value = name_field.get()
                sheet.cell(row=current_row + 1, column=2).value = student_id_field.get()
                sheet.cell(row=current_row + 1, column=3).value = grade_field.get()
                sheet.cell(row=current_row + 1, column=4).value = section_field.get()
                sheet.cell(row=current_row + 1, column=5).value = phone_field.get()
                sheet.cell(row=current_row + 1, column=6).value = email_field.get()
                sheet.cell(row=current_row + 1, column=7).value = address_field.get()

                wb.save(sheet_path)
                name_field.focus_set()
                clear_form()
            delay()

        #Start with sheet initializing
        excel()

        #Submit button
        submit = Button(new_window, text="Submit", fg="Black", bg="light blue", command=insert)
        submit.place(x = 600, y = 550)

        #Frames
        name_frame = Frame(new_window, bg="#262626")
        name_frame.place(x=formx, y=150)

        student_id_frame = Frame(new_window, bg="#262626")
        student_id_frame.place(x=formx, y=200)

        grade_frame = Frame(new_window, bg="#262626")
        grade_frame.place(x=formx, y=250)

        section_frame = Frame(new_window, bg="#262626")
        section_frame.place(x=formx, y=300)

        phone_frame = Frame(new_window, bg="#262626")
        phone_frame.place(x=formx, y=350)

        email_frame = Frame(new_window, bg="#262626")
        email_frame.place(x=formx, y=400)

        address_frame = Frame(new_window, bg="#262626")
        address_frame.place(x=formx, y=450)

        #Labels & Entries
        name_label = Label(name_frame, text="Name               ", bg="#262626", fg="#ffff99", font=("Calibri", 17))
        name_field = Entry(name_frame)
        name_label.grid(row=0, column=0)
        name_field.grid(row=0, column=1, ipadx="100")

        student_id_label = Label(student_id_frame, text="Student ID      ", bg="#262626", fg="#ffff99", font=("Calibri", 17))
        student_id_field = Entry(student_id_frame)
        student_id_label.grid(row=0, column=0)
        student_id_field.grid(row=0, column=1, ipadx="100")

        grade_label = Label(grade_frame, text="Grade               ", bg="#262626", fg="#ffff99", font=("Calibri", 17))
        grade_field = Entry(grade_frame)
        grade_label.grid(row=0, column=0)
        grade_field.grid(row=0, column=1, ipadx="100")

        section_label = Label(section_frame, text="Section            ", bg="#262626", fg="#ffff99", font=("Calibri", 17))
        section_field = Entry(section_frame)
        section_label.grid(row=0, column=0)
        section_field.grid(row=0, column=1, ipadx="100")

        phone_label = Label(phone_frame, text="Phone              ", bg="#262626", fg="#ffff99", font=("Calibri", 17))
        phone_field = Entry(phone_frame)
        phone_label.grid(row=0, column=0)
        phone_field.grid(row=0, column=1, ipadx="100")

        email_label = Label(email_frame, text="Email                ", bg="#262626", fg="#ffff99", font=("Calibri", 17))
        email_field = Entry(email_frame)
        email_label.grid(row=0, column=0)
        email_field.grid(row=0, column=1, ipadx="100")

        address_label = Label(address_frame, text="Address           ", bg="#262626", fg="#ffff99", font=("Calibri", 17))
        address_field = Entry(address_frame)
        address_label.grid(row=0, column=0)
        address_field.grid(row=0, column=1, ipadx="100")

        #End
        return_button()
    sound()
    form_inner()

#Applications
def math():
    def math_inner():
        design()
        logo()

        #Start
        path = tk.Label(new_window, text="MAIN → Applications → Math", bg="#5be471")
        path.place(x = pathx, y = pathy)

        #Math code
        heading = tk.Label(new_window, text="Math", bg="#262626", fg="#99ccff", font=("Calibri", 25))
        heading.place(x = 595, y = pathy)

        #Buttons Color execution functions
        def calculator_enter(nul):
            calculator_button.config(bg="#2c7d88", width=32, height=3)

        def calculator_leave(nul):
            calculator_button.config(bg="SystemButtonFace", width=30, height=2)

        def polynomials_enter(nul):
            polynomials_button.config(bg="#2c7d88", width=32, height=3)

        def polynomials_leave(nul):
            polynomials_button.config(bg="SystemButtonFace", width=30, height=2)

        def conversions_enter(nul):
            conversions_button.config(bg="#2c7d88", width=32, height=3)

        def conversions_leave(nul):
            conversions_button.config(bg="SystemButtonFace", width=30, height=2)

        #Calculator button
        calculator_button = tk.Button(new_window, text = "Calculator", command = calculator, width = 30, height = 2, font="Consolas", activebackground="pink")
        calculator_button.place(x = for3x1, y = for3y)

        #Polynomials button
        polynomials_button = tk.Button(new_window, text = "Polynomials", command = polynomials, width = 30, height = 2, font = "Consolas", activebackground="pink")
        polynomials_button.place(x = for3x2, y = for3y)

        #Conversions button
        conversions_button = tk.Button(new_window, text = "Digital Storage Converter", command = conversions, width = 30, height = 2, font="Consolas", activebackground="pink")
        conversions_button.place(x = for3x3, y = for3y)

        #Buttons Color execution
        calculator_button.bind("<Enter>", calculator_enter)
        calculator_button.bind("<Leave>", calculator_leave)

        polynomials_button.bind("<Enter>", polynomials_enter)
        polynomials_button.bind("<Leave>", polynomials_leave)

        conversions_button.bind("<Enter>", conversions_enter)
        conversions_button.bind("<Leave>", conversions_leave)
        
        #End
        return_button()
    sound()
    math_inner()

def time():
    def time_inner():
        design()
        logo()

        #Start
        path = tk.Label(new_window, text="MAIN → Applications → Time", bg="#5be471")
        path.place(x = pathx, y = pathy)

        #Time code
        heading = tk.Label(new_window, text="Time", bg="#262626", fg="#99ccff", font=("Calibri", 25))
        heading.place(x = 598, y = pathy)

        #Buttons Color execution functions
        def clock_enter(nul):
            clock_button.config(bg="#5bc9d6", width=32, height=3)

        def clock_leave(nul):
            clock_button.config(bg="SystemButtonFace", width=30, height=2)

        def timer_enter(nul):
            timer_button.config(bg="#5bc9d6", width=32, height=3)

        def timer_leave(nul):
            timer_button.config(bg="SystemButtonFace", width=30, height=2)

        def stopwatch_enter(nul):
            stopwatch_button.config(bg="#5bc9d6", width=32, height=3)

        def stopwatch_leave(nul):
            stopwatch_button.config(bg="SystemButtonFace", width=30, height=2)

        #Clock button
        clock_button = tk.Button(new_window, text = "Clock", command = clock, width = 30, height = 2, font="Consolas", activebackground="pink")
        clock_button.place(x = for3x1, y = for3y)

        #Timer button
        timer_button = tk.Button(new_window, text = "Timer", command = timer, width = 30, height = 2, font="Consolas", activebackground="pink")
        timer_button.place(x = for3x2, y = for3y)

        #Stopwatch button
        stopwatch_button = tk.Button(new_window, text = "Stopwatch", command = stopwatch, width = 30, height = 2, font="Consolas", activebackground="pink")
        stopwatch_button.place(x=for3x3, y=for3y)
        
        #Buttons Color execution
        clock_button.bind("<Enter>", clock_enter)
        clock_button.bind("<Leave>", clock_leave)

        timer_button.bind("<Enter>", timer_enter)
        timer_button.bind("<Leave>", timer_leave)

        stopwatch_button.bind("<Enter>", stopwatch_enter)
        stopwatch_button.bind("<Leave>", stopwatch_leave)

        #End
        return_button()
    sound()
    time_inner()

def other():
    def other_inner():
        design()
        logo()

        #Start
        path = tk.Label(new_window, text="MAIN → Applications → Other", bg="#5be471")
        path.place(x = pathx, y = pathy)

        #Other code
        heading = tk.Label(new_window, text="Other", bg="#262626", fg="#99ccff", font=("Calibri", 25))
        heading.place(x = 593, y = pathy)

        #Buttons Color execution functions
        def notepad_enter(nul):
            notepad_button.config(bg="#f3de63", width=32, height=3)

        def notepad_leave(nul):
            notepad_button.config(bg="SystemButtonFace", width=30, height=2)

        def form_enter(nul):
            form_button.config(bg="#f3de63", width=32, height=3)

        def form_leave(nul):
            form_button.config(bg="SystemButtonFace", width=30, height=2)

        #Notepad button
        notepad_button = tk.Button(new_window, text = "Notepad", command = notepad, width = 30, height = 2, font="Consolas", activebackground="pink")
        notepad_button.place(x=250, y=350)
        
        form_button = tk.Button(new_window, text = "Student Form", command = form, width = 30, height = 2, font="Consolas", activebackground="pink")
        form_button.place(x=750, y=350)

        #Buttons Color execution
        notepad_button.bind("<Enter>", notepad_enter)
        notepad_button.bind("<Leave>", notepad_leave)

        form_button.bind("<Enter>", form_enter)
        form_button.bind("<Leave>", form_leave)

        #End
        return_button()
    sound()
    other_inner()

#Main Page
def docs():
    def docs_inner():
        design()
        logo()

        #Start
        path = tk.Label(new_window, text="MAIN → Documentation", bg="#5be471")
        path.place(x = pathx, y = pathy)

        #Docs code
        heading = tk.Label(new_window, text="Documentation", bg="#262626", fg="#99ccff", font=("Calibri", 25))
        heading.place(x = 526, y = pathy)

        #Heading
        text = tk.Label(new_window, text="Project Statistics", height=1, width=20, font=("Calibri", 22), bg="#262626", fg="#1ce2e2")
        text.place(x = for3x1 + 390, y = for3y - 230)

        #Stats headings
        text = tk.Label(new_window,
                         text=f"Project duration\n{i}lines\n{i}pages\n{i}buttons\n{i}color codes\n{i}entry boxes\n{i}frames\n{i}grids\n{i}labels\n{i}functions\n{i}conditions\n{i}libraries imported\n{i}validations",
                           height=20, width=40, font=("Arial", 14), anchor="center", bg="#151516", fg="#e90eb9",
                            justify="left")

        text.place(x = for3x1 + 180, y = for3y - 160)

        #Stats values
        text = tk.Label(new_window,
                         text=f"2.5 months\n2,300+\n18\n67\n242\n25\n8\n98\n80\n117\n95\n10\n141",
                           height=20, width=30, font=("Arial", 14), anchor="center", bg="#151516", fg="#1ad41a",
                            justify="left")

        text.place(x = for3x1 + 580, y = for3y - 160)

        #End
        return_button()
    sound()
    docs_inner()

def applications():
    def applications_inner():
        design()
        logo()

        #Start
        path = tk.Label(new_window, text="MAIN → Applications", bg="#5be471")
        path.place(x = pathx, y = pathy)

        #Applications code
        heading = tk.Label(new_window, text="Applications", bg="#262626", fg="#99ccff", font=("Calibri", 25))
        heading.place(x = 534, y = pathy)

        #Buttons Color execution functions
        def math_enter(nul):
            math_button.config(bg="#b06188", width=32, height=3)

        def math_leave(nul):
            math_button.config(bg="SystemButtonFace", width=30, height=2)

        def time_enter(nul):
            time_button.config(bg="#b06188", width=32, height=3)

        def time_leave(nul):
            time_button.config(bg="SystemButtonFace", width=30, height=2)

        def other_enter(nul):
            other_button.config(bg="#b06188", width=32, height=3)

        def other_leave(nul):
            other_button.config(bg="SystemButtonFace", width=30, height=2)

        #Math button
        math_button = tk.Button(new_window, text = "Math", command = math, width = 30, height = 2, font = "Consolas", activebackground="pink",)
        math_button.place(x = for3x1, y = for3y)

        #Time button
        time_button = tk.Button(new_window, text = "Time", command = time, width = 30, height = 2, font = "Consolas", activebackground="pink")
        time_button.place(x = for3x2, y = for3y)

        #Other button
        other_button = tk.Button(new_window, text = "Other", command = other, width = 30, height = 2, font = "Consolas", activebackground="pink")
        other_button.place(x = for3x3, y = for3y)

        #Buttons Color execution
        math_button.bind("<Enter>", math_enter)
        math_button.bind("<Leave>", math_leave)

        time_button.bind("<Enter>", time_enter)
        time_button.bind("<Leave>", time_leave)

        other_button.bind("<Enter>", other_enter)
        other_button.bind("<Leave>", other_leave)

        #End
        return_button()
    sound()
    applications_inner()

def credits():
    def credits_inner():
        design()
        logo()

        #Start
        path = tk.Label(new_window, text="MAIN → Credits", bg="#5be471")
        path.place(x = pathx, y = pathy)

        #Credits code
        heading = tk.Label(new_window, text="Credits", bg="#262626", fg="#99ccff", font=("Calibri", 25))
        heading.place(x = 583, y = pathy)

        #Credits Titles
        text = tk.Label(new_window, text="Shariq", height=1, width=10, font=("Calibri", 25), bg="#262626", fg="#1ce2e2")
        text.place(x = for3x1 + 55, y = for3y - 200)

        text = tk.Label(new_window, text="Arub", height=1, width=10, font=("Calibri", 25), bg="#262626", fg="#1ce2e2")
        text.place(x = for3x2 + 55, y = for3y - 200)

        text = tk.Label(new_window, text="Ahmed", height=1, width=10, font=("Calibri", 25), bg="#262626", fg="#1ce2e2")
        text.place(x = for3x3 + 55, y = for3y - 200)

        #Credits Content
        text = tk.Label(new_window,
                         text="\nCalculator\nQuadratic\nCubic\nQuartic\nConversions\nClock\nTimer\nStopwatch\n\nFrames\nGrids\nValidations\nNests & Logic",
                           height=18, width=30,
                             font=("Arial", 14), anchor="center", bg="#151516", fg="#1ad41a",
                              justify="left")
        text.place(x = for3x1 - 20, y = for3y - 130)

        text = tk.Label(new_window,
                         text="\nStudent Form\nNotepad\nStructure\nPages\nButtons\nSounds\nBackgrounds\nDesign\nXLSX link\nFunctions\nIncased modules\nPaths\nLabels",
                           height=18, width=30,
                            font=("Arial", 14), anchor="center", bg="#151516", fg="#1ad41a",
                             justify="left")
        text.place(x = for3x2 - 20, y = for3y - 130)

        text = tk.Label(new_window,
                         text="\nStructure Diagram\nFlowchart\nDocs\nCredits\nChart\nPresentation\nProject research\nTesting\nFlyers",
                           height=18, width=30,
                            font=("Arial", 14), anchor="center", bg="#151516", fg="#1ad41a",
                             justify="left")
        text.place(x = for3x3 - 20, y = for3y - 130)

        #Credits Headings
        #Shariq - Modules
        text = tk.Label(new_window, text="Modules", height=1, width=10, font=("Arial", 14), bg="#151516", fg="#e90eb9")
        text.place(x = for3x1 + 39, y = for3y - 85)

        #Shariq - Core Skills
        text = tk.Label(new_window, text="Core Skills", height=1, width=10, font=("Arial", 14), bg="#151516", fg="#e90eb9")
        text.place(x = for3x1 + 45, y = for3y + 112)
        
        #Arub - Modules
        text = tk.Label(new_window, text="Modules", height=1, width=10, font=("Arial", 14), bg="#151516", fg="#e90eb9")
        text.place(x = for3x1 + 435, y = for3y - 87)
        
        #Arub - Structure
        text = tk.Label(new_window, text="Structure", height=1, width=10, font=("Arial", 14), bg="#151516", fg="#e90eb9")
        text.place(x = for3x1 + 435, y = for3y - 20)

        #Ahmed - Representation
        text = tk.Label(new_window, text="Representation", height=1, width=15, font=("Arial", 14), bg="#151516", fg="#e90eb9")
        text.place(x = for3x1 + 820, y = for3y - 43)

        #End
        return_button()
    sound()
    credits_inner()

#Window & BG - MAIN PAGE
window = tk.Tk()
window.attributes('-fullscreen', True)
window.configure(bg = 'black')

#Background
background = PhotoImage(file = wallpaper_path)
background_label = tk.Label(window, image = background)
background_label.place(x = -330, y = -300)

#Buttons on Main Page
def main():

    #Buttons Color execution functions
    def docs_enter(nul):
        docs_button.config(bg="#c5861e", width=32, height=3)

    def docs_leave(nul):
        docs_button.config(bg="SystemButtonFace", width=30, height=2)

    def applications_enter(nul):
        applications_button.config(bg="#24ac72", width=32, height=3)

    def applications_leave(nul):
        applications_button.config(bg="SystemButtonFace", width=30, height=2)

    def credits_enter(nul):
        credits_button.config(bg="#992da3", width=32, height=3)

    def credits_leave(nul):
        credits_button.config(bg="SystemButtonFace", width=30, height=2)


    #Docs button
    docs_button = tk.Button(window, text = "Documentation", command = docs, width = 30, height = 2, font = "Consolas", activebackground="pink")
    docs_button.place(x = for3x1, y = for3y)

    #Applications button
    applications_button = tk.Button(window, text = "Applications", command = applications, width = 30, height = 2, font = "Consolas", activebackground="pink")
    applications_button.place(x = for3x2, y = for3y)

    #Credits button
    credits_button = tk.Button(window, text = "Credits", command = credits, width = 30, height = 2, font = "Consolas", activebackground="pink")
    credits_button.place(x = for3x3, y = for3y)

    #Buttons Color execution
    docs_button.bind("<Enter>", docs_enter)
    docs_button.bind("<Leave>", docs_leave)

    applications_button.bind("<Enter>", applications_enter)
    applications_button.bind("<Leave>", applications_leave)

    credits_button.bind("<Enter>", credits_enter)
    credits_button.bind("<Leave>", credits_leave)

    #Exit button
    #Buttons Color execution functions
    def exit_enter(nul):
        exit_button.config(bg="#e05122", width=13, height=1)

    def exit_leave(nul):
        exit_button.config(bg="SystemButtonFace", width=10, height=1)

    #Main exit button
    exit_button = tk.Button(window, text = "Exit", command=window.destroy, activebackground="red", width=7, height=1, font = ("Consolas", 10))
    exit_button.place(x = 1165, y = 660)

    #Buttons Color execution
    exit_button.bind("<Enter>", exit_enter)
    exit_button.bind("<Leave>", exit_leave)

#Run
main()
window.mainloop()